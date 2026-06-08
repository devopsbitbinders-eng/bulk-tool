"""
=============================================================
  EXCEL CAMPAIGN RESTORE SCRIPT
  Sirf 21, 25, 26 May wale Excel files ke liye
  *** YE SCRIPT KISI KO BHI WHATSAPP MESSAGE NAHI BHEJEGA ***
=============================================================
"""

import asyncio
import json
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

from databases import Database
from dotenv import load_dotenv

load_dotenv()

USER_ID = 1

CAMPAIGNS_TO_RESTORE = [
    {
        "name": "Campaign 21 May",
        "date": "2026-05-21 10:00:00",
        "file": r"C:\Users\kajal\Downloads\campaign_135_report (3).xlsx",
    },
    {
        "name": "Campaign 25 May",
        "date": "2026-05-25 10:00:00",
        "file": r"C:\Users\kajal\Downloads\campaign_139_report.xlsx",
    },
    {
        "name": "Campaign 26 May",
        "date": "2026-05-26 10:00:00",
        "file": r"C:\Users\kajal\Downloads\campaign_141_report.xlsx",
    },
]

def normalize_phone(phone_raw):
    if not phone_raw:
        return None
    phone = str(phone_raw).strip().replace(" ", "").replace("-", "").replace("+", "")
    if phone.endswith(".0"):
        phone = phone[:-2]
    if not phone or not phone.isdigit():
        return None
    if len(phone) == 10:
        phone = "91" + phone
    return phone

def read_excel_file(filepath):
    """Excel file padho aur phone numbers nikalo"""
    import openpyxl
    rows = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Headers dhundo
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')
        
        print(f"   Columns found: {headers}")
        
        # Phone column dhundo
        phone_col_idx = None
        for i, h in enumerate(headers):
            if any(word in h for word in ['phone', 'mobile', 'number', 'contact', 'mo']):
                phone_col_idx = i
                print(f"   Phone column: '{h}' (index {i})")
                break
        
        if phone_col_idx is None:
            print(f"   WARNING: Phone column nahi mila! Pehla column use karunga: '{headers[0]}'")
            phone_col_idx = 0
        
        # Data rows padho
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {headers[i]: str(row[i]) if row[i] is not None else '' for i in range(len(headers)) if i < len(row)}
                phone_raw = row[phone_col_idx] if phone_col_idx < len(row) else None
                rows.append({
                    'phone_raw': phone_raw,
                    'row_data': row_dict
                })
        
        print(f"   Total rows: {len(rows)}")
        wb.close()
        return rows
        
    except Exception as e:
        print(f"   ERROR reading excel: {e}")
        return []

async def restore_excel_campaigns():
    DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///campaigns.db")
    if DATABASE_URL.startswith("mysql://"):
        DATABASE_URL = DATABASE_URL.replace("mysql://", "mysql+aiomysql://", 1)
    if "charset" not in DATABASE_URL and "mysql" in DATABASE_URL:
        DATABASE_URL += "?charset=utf8mb4"
    
    print(f"\nDatabase se connect ho raha hai...")
    db = Database(DATABASE_URL)
    
    try:
        await db.connect()
        print("Connected!\n")
    except Exception as e:
        print(f"ERROR: {e}")
        return
    
    total_campaigns = 0
    total_messages = 0
    
    for i, camp in enumerate(CAMPAIGNS_TO_RESTORE):
        print(f"\n{'='*60}")
        print(f"Campaign {i+1}/{len(CAMPAIGNS_TO_RESTORE)}: {camp['name']}")
        print(f"File: {camp['file']}")
        
        if not os.path.exists(camp['file']):
            print(f"   FILE NAHI MILI! Skip kar raha hun...")
            continue
        
        rows = read_excel_file(camp['file'])
        if not rows:
            print("   Koi data nahi mila, skip...")
            continue
        
        # Valid phones
        valid = []
        for r in rows:
            phone = normalize_phone(r['phone_raw'])
            if phone:
                valid.append({'phone': phone, 'row_data': r['row_data']})
        
        print(f"   Valid phone numbers: {len(valid)}")
        
        if not valid:
            print("   Koi valid number nahi mila, skip...")
            continue
        
        # Campaign insert
        try:
            camp_id = await db.execute("""
                INSERT INTO campaigns 
                (user_id, name, timestamp, total_numbers, sent_success, sent_failed, 
                 status, msg_type, template_name, message_template)
                VALUES 
                (:user_id, :name, :timestamp, :total, :success, :failed, 
                 :status, :msg_type, :template_name, :message_template)
            """, {
                "user_id": USER_ID,
                "name": camp['name'],
                "timestamp": camp['date'],
                "total": len(valid),
                "success": len(valid),
                "failed": 0,
                "status": "completed",
                "msg_type": "template",
                "template_name": "",
                "message_template": "",
            })
            print(f"   Campaign created! ID: {camp_id}")
            total_campaigns += 1
        except Exception as e:
            print(f"   Campaign insert error: {e}")
            continue
        
        # Messages insert
        batch = []
        for r in valid:
            batch.append({
                "user_id": USER_ID,
                "campaign_id": camp_id,
                "phone": r['phone'],
                "status": "delivered",
                "error_message": "",
                "whatsapp_message_id": "",
                "row_data": json.dumps(r['row_data']),
                "timestamp": camp['date'],
            })
        
        inserted = 0
        for j in range(0, len(batch), 100):
            try:
                await db.execute_many("""
                    INSERT INTO messages 
                    (user_id, campaign_id, phone, status, error_message, 
                     whatsapp_message_id, row_data, timestamp)
                    VALUES 
                    (:user_id, :campaign_id, :phone, :status, :error_message,
                     :whatsapp_message_id, :row_data, :timestamp)
                """, batch[j:j+100])
                inserted += len(batch[j:j+100])
                print(f"   Inserting: {inserted}/{len(batch)}...", end='\r')
            except Exception as e:
                print(f"\n   Batch error: {e}")
        
        print(f"\n   {inserted} messages restored!")
        total_messages += inserted
    
    await db.disconnect()
    
    print(f"\n{'='*60}")
    print(f"RESTORE COMPLETE!")
    print(f"Campaigns restored: {total_campaigns}")
    print(f"Messages restored: {total_messages}")
    print(f"Ab Vercel dashboard refresh karo!")
    print(f"*** KISI KO WHATSAPP MESSAGE NAHI GAYA ***")
    print(f"{'='*60}")

if __name__ == "__main__":
    asyncio.run(restore_excel_campaigns())
